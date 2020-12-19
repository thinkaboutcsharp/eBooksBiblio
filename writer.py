def write_biblio(biblio, keys, out_path, out_filename):
    import openpyxl
    import os
    import os.path
    from toutf8 import toutf8

    if len(biblio) == 0:
        return

    out_path = toutf8(out_path)
    out_filename = toutf8(out_filename)
    
    if not os.path.exists(out_path):
        print('output path not exists!')
        return

    out_file_path = os.path.join(out_path, out_filename)
    if os.path.exists(out_file_path):
        os.remove(out_file_path)

    xlb = openpyxl.Workbook()
    xlb.create_sheet('biblio')
    xlb.remove(xlb['Sheet'])

    biblio_sheet = xlb['biblio']
    row = 1
    for book in biblio:
        col = 1
        for key in keys:
            if key in book:
                if isinstance(book[key], list):
                    biblio_sheet.cell(row, col).value = '/'.join(book[key])
                else:
                    biblio_sheet.cell(row, col).value = book[key]
            col = col + 1
        row = row + 1


    xlb.save(out_file_path)


if __name__ == '__main__':
    from reader import read_books
    biblio = read_books('F:\\ScanSnap')
    keys = ['file', 'directory']
    write_biblio(biblio, keys, 'out', 'scan.xlsx')